import streamlit as st
import pandas as pd
import openpyxl
import io
from datetime import datetime

# --- More precise mapping of Excel number formats (No changes here) ---
def get_excel_format_type(number_format):
    """Maps Excel's built-in cell format to a more specific conceptual type."""
    if not number_format:
        return "General"
    fmt = number_format.lower()
    if '_(' in fmt and '*' in fmt and ')' in fmt: return 'Accounting'
    if 'yy' in fmt or 'mm' in fmt or 'dd' in fmt: return 'Date'
    if '%' in fmt: return 'Percentage'
    if any(c in fmt for c in ['$', '€', '£', '¥', '₹']): return 'Currency'
    if '0' in fmt or '#' in fmt: return 'Numeric'
    if fmt == '@': return 'Text'
    if fmt == 'general': return 'General'
    return "Other"

# ---------------- Core Validation Logic (No changes here) ----------------
def compare_excel_files(input_file, output_file):
    results = {}
    try:
        input_wb_fmt = openpyxl.load_workbook(input_file, data_only=False)
        output_wb_fmt = openpyxl.load_workbook(output_file, data_only=False)
        input_wb_val = openpyxl.load_workbook(input_file, data_only=True)
        output_wb_val = openpyxl.load_workbook(output_file, data_only=True)
    except Exception as e:
        st.error(f"Error: Could not load Excel files. Details: {e}")
        return {}

    common_sheets = sorted(set(input_wb_fmt.sheetnames).intersection(output_wb_fmt.sheetnames))
    for sheet_name in common_sheets:
        try:
            results[sheet_name] = []
            ws_in_fmt, ws_out_fmt = input_wb_fmt[sheet_name], output_wb_fmt[sheet_name]
            ws_in_val, ws_out_val = input_wb_val[sheet_name], output_wb_val[sheet_name]
            num_cols, num_rows = ws_in_val.max_column, ws_in_val.max_row
            headers = {c: ws_in_val.cell(row=3, column=c).value for c in range(1, num_cols + 1)}

            # Compare headers (Row 3)
            for col in range(1, num_cols + 1):
                t_val = ws_in_val.cell(row=3, column=col).value
                o_val = ws_out_val.cell(row=3, column=col).value
                if t_val is None or str(t_val).strip() == "": continue
                value_reason = "Values match exactly" if t_val == o_val else f"Template: '{t_val}', Output: '{o_val}'"
                results[sheet_name].append({
                    "SHEET": sheet_name, "CELL": ws_out_val.cell(row=3, column=col).coordinate,
                    "FIELD": headers.get(col), "EXPECTED VALUE": str(t_val), "TEST VALUE": str(o_val),
                    "Data Type Result": "N/A", "Data Type Reason": "Header row - no type check",
                    "Value Result": "Correct" if t_val == o_val else "Wrong",
                    "Value Reason": value_reason
                })

            # Compare data rows (Row 4+)
            for r in range(4, num_rows + 1):
                for c in range(1, num_cols + 1):
                    t_val = ws_in_val.cell(row=r, column=c).value
                    if t_val is None or str(t_val).strip() == "": continue
                    o_val = ws_out_val.cell(row=r, column=c).value
                    t_fmt, o_fmt = ws_in_fmt.cell(row=r, column=c).number_format, ws_out_fmt.cell(row=r, column=c).number_format
                    t_type, o_type = get_excel_format_type(t_fmt), get_excel_format_type(o_fmt)
                    dtype_res = "Correct" if t_type == o_type else "Wrong"
                    dtype_reason = f"Template: '{t_type}', Output: '{o_type}'"
                    val_res = "Correct" if t_val == o_val else "Wrong"
                    val_reason = "Values match exactly" if t_val == o_val else f"Template: '{t_val}', Output: '{o_val}'"
                    results[sheet_name].append({
                        "SHEET": sheet_name, "CELL": ws_out_val.cell(row=r, column=c).coordinate,
                        "FIELD": headers.get(c, f"Col_{c}"), "EXPECTED VALUE": str(t_val), "TEST VALUE": str(o_val),
                        "Data Type Result": dtype_res, "Data Type Reason": dtype_reason,
                        "Value Result": val_res, "Value Reason": val_reason
                    })
        except Exception as e:
            st.warning(f"Error processing sheet '{sheet_name}': {e}")
            continue
    return results

# ---------------- UPDATED: Report Generation with New Dark Header Colors ----------------
def generate_excel_report(results):
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    workbook = writer.book

    all_rows = [row for sheet_rows in results.values() for row in sheet_rows]
    df = pd.DataFrame(all_rows)
    
    # Dashboard calculations (no change)
    data_df = df[df["Data Type Result"] != "N/A"].copy()
    total_checked, dtype_correct, value_correct = len(data_df), len(data_df[data_df["Data Type Result"] == "Correct"]), len(data_df[data_df["Value Result"] == "Correct"])
    dtype_accuracy, value_accuracy = (dtype_correct / total_checked * 100) if total_checked > 0 else 100, (value_correct / total_checked * 100) if total_checked > 0 else 100
    dtype_errors, value_errors = total_checked - dtype_correct, total_checked - value_correct

    # Dashboard sheet (no change)
    dash = workbook.add_worksheet("QA Dashboard")
    title_fmt, kpi_fmt, label_fmt = workbook.add_format({'bold': True, 'font_size': 20, 'align': 'center', 'valign': 'vcenter'}), workbook.add_format({'bold': True, 'font_size': 28, 'align': 'center', 'valign': 'vcenter'}), workbook.add_format({'font_size': 12, 'align': 'center', 'font_color': '#595959'})
    dash.merge_range('B2:G3', 'Validation Dashboard', title_fmt); dash.merge_range('B5:D7', f"{dtype_accuracy:.1f}%", kpi_fmt); dash.merge_range('B8:D8', 'Data Type Accuracy', label_fmt); dash.merge_range('E5:G7', f"{value_accuracy:.1f}%", kpi_fmt); dash.merge_range('E8:G8', 'Value Accuracy', label_fmt); dash.merge_range('B10:D12', dtype_errors, kpi_fmt); dash.merge_range('B13:D13', 'Data Type Errors', label_fmt); dash.merge_range('E10:G12', value_errors, kpi_fmt); dash.merge_range('E13:G13', 'Value Errors', label_fmt); dash.set_column('B:G', 22)

    # --- THIS IS THE CHANGE: Create new formats for dark headers and data cells ---
    cell_wrap_format = workbook.add_format({'text_wrap': True, 'valign': 'top'})
    common_props = {'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'font_color': '#FFFFFF'}
    header_fmt_blue = workbook.add_format({**common_props, 'bg_color': '#002060'})   # Dark Blue
    header_fmt_red = workbook.add_format({**common_props, 'bg_color': '#C00000'})     # Dark Red
    header_fmt_green = workbook.add_format({**common_props, 'bg_color': '#00B050'}) # Dark Green

    # --- Sheet 2: Data Type Results with new color-coded headers ---
    dtype_cols = ["SHEET", "CELL", "FIELD", "Data Type Result", "Data Type Reason"]
    dtype_df = df[df["Data Type Result"] != "N/A"][dtype_cols]
    dtype_df.to_excel(writer, sheet_name="Data Type Results", index=False, header=False, startrow=1)
    ws_dtype = writer.sheets["Data Type Results"]
    for col_num, value in enumerate(dtype_df.columns.values):
        if value in ["SHEET", "CELL", "FIELD"]:
            ws_dtype.write(0, col_num, value, header_fmt_blue)
        else: # Data Type Result, Data Type Reason
            ws_dtype.write(0, col_num, value, header_fmt_green)
    ws_dtype.set_column('A:E', 25, cell_wrap_format)

    # --- Sheet 3: Value Match Results with new color-coded headers ---
    value_cols = ["SHEET", "CELL", "FIELD", "EXPECTED VALUE", "TEST VALUE", "Value Result", "Value Reason"]
    value_df = df[value_cols]
    value_df.to_excel(writer, sheet_name="Value Match Results", index=False, header=False, startrow=1)
    ws_value = writer.sheets["Value Match Results"]
    for col_num, value in enumerate(value_df.columns.values):
        if value in ["SHEET", "CELL", "FIELD"]:
            ws_value.write(0, col_num, value, header_fmt_blue)
        elif value in ["EXPECTED VALUE", "TEST VALUE"]:
            ws_value.write(0, col_num, value, header_fmt_red)
        else: # Value Result, Value Reason
            ws_value.write(0, col_num, value, header_fmt_green)
    ws_value.set_column('A:G', 25, cell_wrap_format)

    writer.close()
    output.seek(0)
    return output

# ---------------- Streamlit UI (No changes here) ----------------
st.set_page_config(page_title="Excel Validator", layout="wide")
st.title("Excel Validator — Data Type & Value Check")

def reset_state():
    st.session_state.ran = False
    st.session_state.results = {}

if 'ran' not in st.session_state: st.session_state.ran = False
if 'results' not in st.session_state: st.session_state.results = {}

input_file = st.file_uploader("Upload Input Template Excel", type=['xlsx'])
output_file = st.file_uploader("Upload Output Excel to Test", type=['xlsx'])

if input_file and output_file:
    if st.button("Run Validation", type="primary"):
        with st.spinner("Validating..."):
            st.session_state.results = compare_excel_files(input_file, output_file)
        st.session_state.ran = True

if st.session_state.ran:
    res = st.session_state.results
    if res:
        all_rows = [row for sheet_rows in res.values() for row in sheet_rows]
        df = pd.DataFrame(all_rows)
        data_df = df[df["Data Type Result"] != "N/A"].copy()
        total_checked = len(data_df)
        dtype_correct = len(data_df[data_df["Data Type Result"] == "Correct"])
        value_correct = len(data_df[data_df["Value Result"] == "Correct"])
        dtype_accuracy = (dtype_correct / total_checked * 100) if total_checked > 0 else 100
        value_accuracy = (value_correct / total_checked * 100) if total_checked > 0 else 100
        dtype_errors = total_checked - dtype_correct
        value_errors = total_checked - value_correct

        st.header("Validation Summary")
        col1, col2 = st.columns(2)
        col1.metric("📊 Data Type Accuracy", f"{dtype_accuracy:.1f}%", f"{dtype_errors} Errors", delta_color="inverse")
        col2.metric("🔢 Value Accuracy", f"{value_accuracy:.1f}%", f"{value_errors} Errors", delta_color="inverse")

        st.download_button(
            "📄 Download Full Test Report",
            data=generate_excel_report(res),
            file_name=f"Test_Report_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            on_click=reset_state
        )

